from ast import Store
from asyncore import loop
from cmath import nan
from gc import collect
from importlib.metadata import files
from itertools import count
from multiprocessing.sharedctypes import Value
from re import U
import string
from turtle import color
from unicodedata import name
import pandas as pd
import openpyxl 
from pandas import ExcelWriter
import glob
from openpyxl.utils import rows_from_range
from openpyxl.styles import Alignment, PatternFill
import tkinter as tk
from tkinter import filedialog
from array import *

#Inputs..............................................................................................................................
path = "C:/Users/bikal/Downloads/20th Street_AWSC_TWSC_Sig_FInal_220720"
filenames = glob.glob(path + "\*.txt")


#excel works................................................................................................................
wb = openpyxl.Workbook()
wb.save('C:/Users/bikal/Downloads/support.xlsx')
path1= r'C:/Users/bikal/Downloads/support.xlsx'
book = openpyxl.load_workbook(path1)
sheet=book.active

#Variable declaration..............................................................................................................................
data_store=[" "]*12
check_na=[0]*len(data_store)
dirn=["EBL","EBT","EBR","WBL","WBT","WBR","NBL","NBT","NBR","SBL","SBT","SBR"]
loop1=0
global loop2
loop2=0
orderm=[0]*(100)
street_name=[0]*15
data_sort=pd.DataFrame()





#excel writer....................................................................................
def writer(data_output):
    #global data_output
    writer = pd.ExcelWriter(path1,engine='openpyxl')
    writer.book = book
    writer.sheets= dict((ws.title,ws) for ws in book.worksheets)
    data_output.to_excel(writer,sheet_name='Sheet',header=None,index=None,startcol=0,startrow=sheet.max_row)
    writer.save()

#unsignalized TWSC...................................................................................................................................
def unsignalized_twsc_awsc(data,check_str,unsig_name,name1):
    global loop1
    global loop2
    global intersection
    global st_name
    global data_output
    for i in range(len(data)):
        string= data.iloc[i,0]
        if string[0:8]=="Movement":
            mo=i
        if string[0:19]=="Lane Configurations":
            co=i
        if string[0:9]=="Int Delay":
            if str(data.iloc[i,2])==str(0):
                break
        if string[0:12]=="HCM Lane LOS":
            loop1=loop1+1
            for j in range(len(data.columns)):
                if check_str[0:4]!='AWSC':
                    lane_t=str(data.iloc[i-4,j])
                else:
                    lane_t=str(data.iloc[i-18,j])
                if lane_t[-3:]=="Ln1":
                    var=lane_t[0:2]
                    for k in range(len(data.columns)):
                        lane_configuration_t=str(data.iloc[co,k])
                        if lane_configuration_t=="<1>" or "<1" or "1>":
                            movement_t=str(data.iloc[co-1,k])
                            if movement_t[0:2]==var:
                                for l in range(len(dirn)):
                                    if movement_t[0:3]==dirn[l]:
                                        if data.iloc[i,j]=="-" or data.iloc[i-1,j]=="-":
                                            data_store[l]="-"
                                        else:
                                            data_store[l]=data.iloc[i+1,j]
                for l in range(len(dirn)):
                    if lane_t[0:3]==dirn[l]:
                        if lane_t[-3:]!="Ln1":
                            if data.iloc[i,j]=="-" or data.iloc[i-1,j]=="-":
                                data_store[l]="-"
                            else:
                                data_store[l]=data.iloc[i+1,j]
            for p in range(len(data.columns)):
                for l in range(len(dirn)):
                    if check_str[0:4]!='AWSC':
                        if str(data.iloc[i-4,p])==dirn[l] or str(data.iloc[mo,p])==dirn[l]:
                            check_na[l]=check_na[l]+1
                    else:
                        if str(data.iloc[i-18,p])==dirn[l] or str(data.iloc[mo,p])==dirn[l]:
                            check_na[l]=check_na[l]+1
            for l in range(len(dirn)):
                if check_na[l]>0:
                    pass
                else:
                    data_store[l]="N/A"
            for l in range(len(check_na)):
                check_na[l]=0
            if loop1==1:
                st_name= unsig_name
            else:
                st_name=data.iloc[i-42,0] 
            alternatives=name1.split("-")[len(name1.split("-"))-1][:-18]         
            data_output=data_output.append({'S/U_Intersection':st_name,'Intersection':alternatives[1:],'EB-Left':data_store[0],'EB-Thru':data_store[1],'EB-Right':data_store[2],'WB-Left':data_store[3],'WB-Thru':data_store[4],'WB-Right':data_store[5],'NB-Left':data_store[6],'NB-Thru':data_store[7],'NB-Right':data_store[8],'SB-Left':data_store[9],'SB-Thru':data_store[10],'SB-Right':data_store[11]},ignore_index=True)
            for l in range(len(data_store)):
                data_store[l]=""
    #data_output.sort_values(by=["S/U_Intersection","Intersection","EB-Left","EB-Thru","EB-Right","WB-Left","WB-Thru","WB-Right","NB-Left","NB-Thru","NB-Right","SB-Left","SB-Thru","SB-Right"],ascending=True)
    #writer(data_output) 
    loop1=0
    st_name=""

#signalized...................................................................................................................
def signalized(data,sig_name,name2):
    global loop2
    global data_output
    for i in range(len(data)):
        string=str(data.iloc[i,0])
        if string[0:17]=="Queue Length 95th":
            loop2=loop2+1
            for j in range(len(data.columns)):
                string1=str(data.iloc[i,j])
                if string1!= "NaN":
                    if string1!="nan":
                        if loop2==1:
                            st_name= sig_name
                        else:
                            st_name=data.iloc[i-55,0]
                        data_store[j-2]=str(data.iloc[i,j])
            alternatives=name2.split("-")[len(name2.split("-"))-1][:-18]
            for l in range(len(data_store)):
                if len(data_store[l])>10:
                    data_store[l]=" "
            data_output=data_output.append({'S/U_Intersection':st_name,'Intersection':alternatives[1:],'EB-Left':data_store[0],'EB-Thru':data_store[1],'EB-Right':data_store[2],'WB-Left':data_store[3],'WB-Thru':data_store[4],'WB-Right':data_store[5],'NB-Left':data_store[6],'NB-Thru':data_store[7],'NB-Right':data_store[8],'SB-Left':data_store[9],'SB-Thru':data_store[10],'SB-Right':data_store[11]},ignore_index=True)
            for l in range(len(data_store)):
                data_store[l]=""
    loop2=0
    #data_output.sort_values(by=["S/U_Intersection","Intersection","EB-Left","EB-Thru","EB-Right","WB-Left","WB-Thru","WB-Right","NB-Left","NB-Thru","NB-Right","SB-Left","SB-Thru","SB-Right"],ascending=True)
    #writer(data_output) 
    st_name=""


#check for signalized, unsignalized....................................................................................................
def check():
    global data_output
    k=0
    counts=0
    orderm=order()
    c=0
    data_output= pd.DataFrame(columns=['S/U_Intersection','Intersection','EB-Left','EB-Thru','EB-Right','WB-Left','WB-Thru','WB-Right','NB-Left','NB-Thru','NB-Right','SB-Left','SB-Thru','SB-Right'])
    for file in range(len(filenames)):
        data=pd.read_csv(orderm[file],sep="\t",header=3,skip_blank_lines=True)
        data1=pd.read_csv(orderm[file-1],sep="\t",header=3,skip_blank_lines=True)
        text_file_raw=open(orderm[file],"r")
        text_file=text_file_raw.read()
        check_str=text_file.split(" ")[2]
        unsig_name=text_file[14:35]
        sig_name=text_file[23:49]
        for i in range(len(data)):
            string=str(data.iloc[i,0])
            if string[0:12]=="HCM Lane LOS":
                k=k+1
                break
        for j in range(len(data1)):
            string1=str(data1.iloc[j,0])
            if string1[0:12]=="HCM Lane LOS":
                c=c+1
                break
        condition1=orderm[file-1].split("-")[len(orderm[file-1].split("-"))-1][:-11]
        splits1=condition1.split(" ")
        time1=splits1[len(splits1)-2]
        condition2=orderm[file].split("-")[len(orderm[file].split("-"))-1][:-11]
        splits2=condition2.split(" ")
        time2=splits2[len(splits2)-2]
        if time1!=time2:
            counts=0
        else:
            counts=counts+1
        if k!=c:
            counts=0
        if splits2[0]!="I":
            if counts==0:
                if len(data_output)==0:
                    table(0,peak_hour=time2+" Peak Hour")
                    pass
                else:
                    data_sort=data_output.sort_values(by='S/U_Intersection')
                    sorted_data=data_frame_management(data_sort)
                    writer(sorted_data)
                    data_output= pd.DataFrame(columns=['S/U_Intersection','Intersection','EB-Left','EB-Thru','EB-Right','WB-Left','WB-Thru','WB-Right','NB-Left','NB-Thru','NB-Right','SB-Left','SB-Thru','SB-Right'])
                    table(0,peak_hour=time2+" Peak Hour")
                    unsignalized_twsc_awsc(data,check_str,unsig_name,name1=orderm[file])
            else:
                unsignalized_twsc_awsc(data,check_str,unsig_name,name1=orderm[file])
        else:
            if counts==0:
                data_sort=data_output.sort_values(by='S/U_Intersection')
                sorted_data=data_frame_management(data_sort)
                writer(sorted_data)
                data_output= pd.DataFrame(columns=['S/U_Intersection','Intersection','EB-Left','EB-Thru','EB-Right','WB-Left','WB-Thru','WB-Right','NB-Left','NB-Thru','NB-Right','SB-Left','SB-Thru','SB-Right'])
                table(1,time2+" Peak Hour")
                signalized(data,sig_name,name2=orderm[file])
            else:
                signalized(data,sig_name,name2=orderm[file])
        k=0
        c=0
    data_sort=data_output.sort_values(by='S/U_Intersection')
    sorted_data=data_frame_management(data_sort)
    writer(sorted_data)    

#.......................................................................................
def data_frame_management(input_data):
    if len(input_data)==0:
        pass
    else:
        for i in range(len(input_data)):
            if input_data.iloc[i,0]!=" ":
                for j in range(len(input_data)):
                    if i!=j:
                        if input_data.iloc[i,0]==input_data.iloc[j,0]:
                            input_data.iloc[j,0]=" "
        return input_data
            

#......................................................................................................
# def table_management():
#     row_value=1
#     if row_value < sheet.max_row:
#         check_cell=sheet.cell(row=row_value,column=1)
#         if check_cell.value!=" ":
#             Value=check_cell.value
#             check_cell_2=sheet.cell(row=row_value+1,column=1)
#             if check_cell_2.value==" ":
#                 sheet.merge_cells(check_cell,check_cell_2)
#                 cell = sheet.cell(row=row_value, column=1)
#                 cell.value=Value  
#             else:
#                 row_value=row_value+1
#         else:
#             row_value=row_value+1
                








#put files in order........................................................................................................................
def order():
    u=0
    tot_time=0
    global am_arr
    am_arr=[0]*len(filenames)
    global pm_arr
    pm_arr=[0]*len(filenames)
    len_am=0
    len_pm=0
    for file in filenames:
        condition=file.split("-")[len(file.split("-"))-1][:-11]
        splits=condition.split(" ")
        time=splits[len(splits)-2]
        if time=="AM":
            am_arr[tot_time]=file
            len_am=len_am+1
        else:
            pm_arr[tot_time]=file
            len_pm=len_pm+1
        tot_time=tot_time+1
    filtered_am_arr=[x for x in am_arr if x!=0]
    filtered_pm_arr=[x for x in pm_arr if x!=0]
    for i in range(len(filtered_am_arr)):
        sfile=str(filtered_am_arr[i])
        name=sfile.split("-")[len(sfile.split("-"))-1][:-11]
        for s in range(len(filtered_am_arr)):
            count=str(filtered_am_arr[s])
            variable=count.split("-")[len(count.split("-"))-1][:-11]
            if name==variable or "A"+name==variable:
                if "I"+name!=variable:
                    orderm[u]=count
                    u=u+1
    for i in range(len(filtered_am_arr)):
        sfile=str(filtered_am_arr[i])
        name=sfile.split("-")[len(sfile.split("-"))-1][:-11]
        for s in range(len(filtered_am_arr)):
            count=str(filtered_am_arr[s])
            variable=count.split("-")[len(count.split("-"))-1][:-11]
            if "I"+name==variable:
                orderm[u]=count
                u=u+1
    for i in range(len(filtered_pm_arr)):
        sfile=str(filtered_pm_arr[i])
        name=sfile.split("-")[len(sfile.split("-"))-1][:-11]
        for s in range(len(filtered_pm_arr)):
            count=str(filtered_pm_arr[s])
            variable=count.split("-")[len(count.split("-"))-1][:-11]
            if name==variable or "A"+name==variable:
                if "I"+name!=variable:
                    orderm[u]=count
                    u=u+1
    for i in range(len(filtered_am_arr)):
        sfile=str(filtered_am_arr[i])
        name=sfile.split("-")[len(sfile.split("-"))-1][:-11]
        for s in range(len(filtered_am_arr)):
            count=str(filtered_am_arr[s])
            variable=count.split("-")[len(count.split("-"))-1][:-11]
            if "I"+name==variable:
                orderm[u]=count
                u=u+1
    filtered_orderm=[x for x in orderm if x!=0]
    filtered_2orderm=[]
    [filtered_2orderm.append(x) for x in filtered_orderm if x not in filtered_2orderm]
    return(filtered_2orderm)




#table..........................................................................................................................
def table(q,peak_hour):
    r=sheet.max_row
    color = openpyxl.styles.colors.Color(rgb='D3D3D3')

    sheet.merge_cells('A'+str(r+1)+':'+'A'+str(r+3))
    cell1 = sheet.cell(row=r+1, column=1)  
    if q==0:
        cell1.value = 'Unsignalized Intersections'
    else:
        cell1.value = 'Signalized Intersections'

    cell1.alignment = Alignment(horizontal='center', vertical='center')  
    cell1.fill=PatternFill(patternType='solid',fgColor=color)

    sheet.merge_cells('B'+str(r+1)+':'+'N'+str(r+1))
    cell2 = sheet.cell(row=r+1, column=2)  
    cell2.value = peak_hour 
    cell2.alignment = Alignment(horizontal='center', vertical='center')
    cell2.fill=PatternFill(patternType='solid',fgColor=color) 

    sheet.merge_cells('C'+str(r+2)+':'+'E'+str(r+2))
    cell15 = sheet.cell(row=r+2, column=3)  
    cell15.value = 'EB' 
    cell15.alignment = Alignment(horizontal='center', vertical='center')
    cell15.fill=PatternFill(patternType='solid',fgColor=color)

    sheet.merge_cells('F'+str(r+2)+':'+'H'+str(r+2))
    cell3 = sheet.cell(row=r+2, column=6)  
    cell3.value = 'WB' 
    cell3.alignment = Alignment(horizontal='center', vertical='center')
    cell3.fill=PatternFill(patternType='solid',fgColor=color) 

    sheet.merge_cells('I'+str(r+2)+':'+'K'+str(r+2))
    cell4 = sheet.cell(row=r+2, column=9)  
    cell4.value = 'NB' 
    cell4.alignment = Alignment(horizontal='center', vertical='center')
    cell4.fill=PatternFill(patternType='solid',fgColor=color) 

    sheet.merge_cells('L'+str(r+2)+':'+'N'+str(r+2))
    cell5 = sheet.cell(row=r+2, column=12)  
    cell5.value = 'SB' 
    cell5.alignment = Alignment(horizontal='center', vertical='center')
    cell5.fill=PatternFill(patternType='solid',fgColor=color)

    sheet.merge_cells('B'+str(r+2)+':'+'B'+str(r+3))
    cell6 = sheet.cell(row=r+2, column=2)  
    cell6.value = 'Alternatives' 
    cell6.alignment = Alignment(horizontal='center', vertical='center')
    cell6.fill=PatternFill(patternType='solid',fgColor=color)


    sheet['C'+str(r+3)]="Left"
    sheet['C'+str(r+3)].fill=PatternFill(patternType='solid',fgColor=color)
    sheet['D'+str(r+3)]="Thru"
    sheet['D'+str(r+3)].fill=PatternFill(patternType='solid',fgColor=color)
    sheet['E'+str(r+3)]="Right"
    sheet['E'+str(r+3)].fill=PatternFill(patternType='solid',fgColor=color)
    sheet['F'+str(r+3)]="Left"
    sheet['F'+str(r+3)].fill=PatternFill(patternType='solid',fgColor=color)
    sheet['G'+str(r+3)]="Thru"
    sheet['G'+str(r+3)].fill=PatternFill(patternType='solid',fgColor=color)
    sheet['H'+str(r+3)]="Right"
    sheet['H'+str(r+3)].fill=PatternFill(patternType='solid',fgColor=color)
    sheet['I'+str(r+3)]="Left"
    sheet['I'+str(r+3)].fill=PatternFill(patternType='solid',fgColor=color)
    sheet['J'+str(r+3)]="Thru"
    sheet['J'+str(r+3)].fill=PatternFill(patternType='solid',fgColor=color)
    sheet['K'+str(r+3)]="Right"
    sheet['K'+str(r+3)].fill=PatternFill(patternType='solid',fgColor=color)
    sheet['L'+str(r+3)]="Left"
    sheet['L'+str(r+3)].fill=PatternFill(patternType='solid',fgColor=color)
    sheet['M'+str(r+3)]="Thru"
    sheet['M'+str(r+3)].fill=PatternFill(patternType='solid',fgColor=color)
    sheet['N'+str(r+3)]="Right"
    sheet['N'+str(r+3)].fill=PatternFill(patternType='solid',fgColor=color)



#sorting...........................................................................................................................




    
#export excel file.................................................................................................................

def exportCSV ():
    export_file_path = filedialog.asksaveasfilename(defaultextension='.xlsx')
    book.save(export_file_path)


#Output in dialogbox...................................................................................................................
def output():
    root= tk.Tk()
    canvas1 = tk.Canvas(root, width = 300, height = 200, bg = 'lightsteelblue2', relief = 'raised')
    canvas1.pack()

    saveAsButton_CSV = tk.Button(text='Export LOS Table', command=exportCSV, bg='green', fg='white', font=('helvetica', 12, 'bold'))
    canvas1.create_window(150, 100, window=saveAsButton_CSV)

    root.mainloop()
        


#Program...................................................................................................................................................

check()
output()

# def collect_st_name(data,unsig_name,sig_name):
#     u=0
#     for i in range(len(data)):
#         string= data.iloc[i,0]
#         if string[0:12]=="HCM Lane LOS":
#             loop1=loop1+1
#             if loop1==1:
#                 st_name= unsig_name
#             else:
#                 st_name=data.iloc[i-42,0]
#             for i in range(len(street_name)):
#                 if street_name[i][0:2]!=st_name[0:2]:
#                     k=k+1
#             if k>1:
#                 street_name[u]=st_name
#                 u=u+1
#             else:
#                 pass
#         if string[0:17]=="Queue Length 95th":
#             loop2=loop2+1
#             if loop2==1:
#                 st_name= sig_name
#             else:
#                 st_name=data.iloc[i-71,0]
#             for i in range(len(street_name)):
#                 if street_name[i][0:2]!=st_name[0:2]:
#                     k=k+1
#             if k>1:
#                 street_name[u]=st_name
#                 u=u+1
#             else:
#                 pass
        
    