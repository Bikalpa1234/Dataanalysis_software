from ast import Store
from asyncore import loop
from cmath import nan
from importlib.metadata import files
from itertools import count
import string
from turtle import color
from unicodedata import name
import pandas as pd
import openpyxl 
import glob
from openpyxl.utils import rows_from_range
from openpyxl.styles import Alignment, PatternFill
import tkinter as tk
from tkinter import filedialog
from array import *

#Inputs..............................................................................................................................
path = "C:/Users/bikal/Downloads/Intersection +HCM 2000"
filenames = glob.glob(path + "\*.txt")


#excel works................................................................................................................
wb = openpyxl.Workbook()
wb.save('C:/Users/bikal/Downloads/support.xlsx')
path1= r'C:/Users/bikal/Downloads/support.xlsx'
book = openpyxl.load_workbook(path1)
sheet=book.active

#Variable declaration..............................................................................................................................
data_store=[" "]*12
check_na=[0]*len(data_store)
dirn=["EBL","EBT","EBR","WBL","WBT","WBR","NBL","NBT","NBR","SBL","SBT","SBR"]
loop1=0
global loop2
loop2=0
orderm=[0]*(len(filenames)+1)




#excel writer....................................................................................
def writer(data_output):
    #global data_output
    writer = pd.ExcelWriter(path1,engine='openpyxl')
    writer.book = book
    writer.sheets= dict((ws.title,ws) for ws in book.worksheets)
    data_output.to_excel(writer,sheet_name='Sheet',header=None,index=False,startcol=0,startrow=sheet.max_row)
    writer.save()

#unsignalized TWSC...................................................................................................................................
def unsignalized_twsc_awsc(data):
    global loop1
    global loop2
    global intersection
    global st_name
    global data_output
    for i in range(len(data)):
        string= data.iloc[i,0]
        if string[0:8]=="Movement":
            mo=i
        if string[0:5]=="Lanes":
            co=i
        if string[0:8]=="Lane LOS":
            loop1=loop1+1
            if str(data.iloc[i+4,4])=="0.0":
                break
            for j in range(len(data.columns)):
                lane_t=str(data.iloc[i-8,j])
                if lane_t[-2:]==" 1":
                    var=lane_t[0:2]
                    for k in range(len(data.columns)):
                        lane_configuration_t=str(data.iloc[co,k])
                        if lane_configuration_t!=str(0) and lane_configuration_t!="Nan" and lane_configuration_t!="nan":
                            movement_t=str(data.iloc[co-1,k])
                            if movement_t[0:2]==var:
                                for l in range(len(dirn)):
                                    if movement_t[0:3]==dirn[l]:
                                        if data.iloc[i,j]=="nan" or data.iloc[i-1,j]=="0.0":
                                            data_store[l]="-"
                                        else:
                                            data_store[l]=str(data.iloc[i,j])+'/'+str(data.iloc[i-1,j])
                for l in range(len(dirn)):
                    if lane_t[0:3]==dirn[l]:
                        if lane_t[-2:]!=" 1":
                            if data.iloc[i,j]=="-" or data.iloc[i-1,j]=="-":
                                data_store[l]="-"
                            else:
                                data_store[l]=str(data.iloc[i,j])+'/'+str(data.iloc[i-1,j])
            for p in range(len(data.columns)):
                for l in range(len(dirn)):
                    if str(data.iloc[i-8,p])==dirn[l] or str(data.iloc[mo,p])==dirn[l]:
                        check_na[l]=check_na[l]+1
            for l in range(len(dirn)):
                if check_na[l]>0:
                    pass
                else:
                    data_store[l]="N/A"
            for l in range(len(check_na)):
                check_na[l]=0
            st_name=data.iloc[i-35,0]           
            intersection=""
            data_output=data_output.append({'S/U_Intersection':st_name,'Intersection':intersection,'EB-Left':data_store[0],'EB-Thru':data_store[1],'EB-Right':data_store[2],'WB-Left':data_store[3],'WB-Thru':data_store[4],'WB-Right':data_store[5],'NB-Left':data_store[6],'NB-Thru':data_store[7],'NB-Right':data_store[8],'SB-Left':data_store[9],'SB-Thru':data_store[10],'SB-Right':data_store[11]},ignore_index=True)
            for l in range(len(data_store)):
                data_store[l]="" 
    st_name=""

#signalized...................................................................................................................
def signalized(data,sig_name,sig_name_2):
    global loop2
    global data_output
    for i in range(len(data)):
        string=str(data.iloc[i,0])
        if string[0:22]=="HCM 2000 Control Delay":
            loop2=loop2+1
            for j in range(len(data.columns)):
                string1=str(data.iloc[i,j])
                if string1[0:25]=="HCM 2000 Level of Service":
                    print('\n',str(data.iloc[i,4]),'\n')
                    if (loop1==1 and loop2==0)or(loop1==0 and loop2==1):
                        st_name= sig_name
                    else:
                        st_name=data.iloc[i-36,0]
                    data_output=data_output.append({'S/U_Intersection':st_name,'Intersection':str(data.iloc[i,j+5])+'/'+str(data.iloc[i,4]),'EB-Left':data_store[0],'EB-Thru':data_store[1],'EB-Right':data_store[2],'WB-Left':data_store[3],'WB-Thru':data_store[4],'WB-Right':data_store[5],'NB-Left':data_store[6],'NB-Thru':data_store[7],'NB-Right':data_store[8],'SB-Left':data_store[9],'SB-Thru':data_store[10],'SB-Right':data_store[11]},ignore_index=True)
                    for l in range(len(data_store)):
                        data_store[l]=""
        if string[0:25]=="Intersection Signal Delay":
            loop2=loop2+1
            for j in range(len(data.columns)):
                string1=str(data.iloc[i,j])
                if string1[0:16]=="Intersection LOS":
                    if loop2==1:
                        st_name= sig_name_2
                    else:
                        st_name=data.iloc[i-71,0]
                    data_output=data_output.append({'S/U_Intersection':st_name,'Intersection':string1.split(':')[1]+'/'+string.split(':')[1],'EB-Left':data_store[0],'EB-Thru':data_store[1],'EB-Right':data_store[2],'WB-Left':data_store[3],'WB-Thru':data_store[4],'WB-Right':data_store[5],'NB-Left':data_store[6],'NB-Thru':data_store[7],'NB-Right':data_store[8],'SB-Left':data_store[9],'SB-Thru':data_store[10],'SB-Right':data_store[11]},ignore_index=True)
                    for l in range(len(data_store)):
                        data_store[l]=""
        if string[0:16]=="Level of Service" or string[0:3]=="LOS":
            for j in range(2,len(data.columns)):
                string1=str(data.iloc[i,j])
                if string1!= "NaN":
                    if string1!="nan":
                        data_store[j-2]=str(data.iloc[i,j])+'/'+str(data.iloc[i-1,j]) 
    st_name=""


#check for signalized, unsignalized....................................................................................................
def check():
    global data_output
    global loop1
    global loop2
    for file in filenames:
        data_output= pd.DataFrame(columns=['S/U_Intersection','Intersection','EB-Left','EB-Thru','EB-Right','WB-Left','WB-Thru','WB-Right','NB-Left','NB-Thru','NB-Right','SB-Left','SB-Thru','SB-Right'])
        data=pd.read_csv(file,sep="\t",header=3,skip_blank_lines=True)
        text_file_raw=open(file,"r")
        text_file=text_file_raw.read()
        sig_name=text_file[46:76]
        sig_name_2=text_file[23:49]
        unsignalized_twsc_awsc(data)
        if loop1>0:
            table(0,file)
            writer(data_output)
            data_output= pd.DataFrame(columns=['S/U_Intersection','Intersection','EB-Left','EB-Thru','EB-Right','WB-Left','WB-Thru','WB-Right','NB-Left','NB-Thru','NB-Right','SB-Left','SB-Thru','SB-Right'])
            loop1=0
        signalized(data,sig_name=sig_name,sig_name_2=sig_name_2)
        if loop2>0:
            table(1,file)
            writer(data_output)
            data_output= pd.DataFrame(columns=['S/U_Intersection','Intersection','EB-Left','EB-Thru','EB-Right','WB-Left','WB-Thru','WB-Right','NB-Left','NB-Thru','NB-Right','SB-Left','SB-Thru','SB-Right'])
            loop2=0 
                   


#put files in order........................................................................................................................
# def order():
#     u=0
#     for sfile in filenames:
#         name=sfile.split("-")[len(sfile.split("-"))-1][:-11]
#         for count in filenames:
#             variable=count.split("-")[len(count.split("-"))-1][:-11]
#             if name==variable or "A"+name==variable or "I"+name==variable:
#                 if u>len(filenames):
#                     break
#                 else:
#                     orderm[u]=count
#                     u=u+1






#table..........................................................................................................................
def table(q,file):
    r=sheet.max_row
    color = openpyxl.styles.colors.Color(rgb='D3D3D3')

    sheet.merge_cells('A'+str(r+1)+':'+'A'+str(r+3))
    cell1 = sheet.cell(row=r+1, column=1)  
    if q==0:
        cell1.value = 'Unsignalized Intersections'
    else:
        cell1.value = 'Signalized Intersections'

    cell1.alignment = Alignment(horizontal='center', vertical='center')  
    cell1.fill=PatternFill(patternType='solid',fgColor=color)

    sheet.merge_cells('B'+str(r+1)+':'+'N'+str(r+1))
    cell2 = sheet.cell(row=r+1, column=2)  
    cell2.value = file.split("-")[len(file.split("-"))-1][:-11] 
    cell2.alignment = Alignment(horizontal='center', vertical='center')
    cell2.fill=PatternFill(patternType='solid',fgColor=color) 

    sheet.merge_cells('C'+str(r+2)+':'+'E'+str(r+2))
    cell15 = sheet.cell(row=r+2, column=3)  
    cell15.value = 'EB' 
    cell15.alignment = Alignment(horizontal='center', vertical='center')
    cell15.fill=PatternFill(patternType='solid',fgColor=color)

    sheet.merge_cells('F'+str(r+2)+':'+'H'+str(r+2))
    cell3 = sheet.cell(row=r+2, column=6)  
    cell3.value = 'WB' 
    cell3.alignment = Alignment(horizontal='center', vertical='center')
    cell3.fill=PatternFill(patternType='solid',fgColor=color) 

    sheet.merge_cells('I'+str(r+2)+':'+'K'+str(r+2))
    cell4 = sheet.cell(row=r+2, column=9)  
    cell4.value = 'NB' 
    cell4.alignment = Alignment(horizontal='center', vertical='center')
    cell4.fill=PatternFill(patternType='solid',fgColor=color) 

    sheet.merge_cells('L'+str(r+2)+':'+'N'+str(r+2))
    cell5 = sheet.cell(row=r+2, column=12)  
    cell5.value = 'SB' 
    cell5.alignment = Alignment(horizontal='center', vertical='center')
    cell5.fill=PatternFill(patternType='solid',fgColor=color)

    sheet.merge_cells('B'+str(r+2)+':'+'B'+str(r+3))
    cell6 = sheet.cell(row=r+2, column=2)  
    cell6.value = 'Intersection' 
    cell6.alignment = Alignment(horizontal='center', vertical='center')
    cell6.fill=PatternFill(patternType='solid',fgColor=color)


    sheet['C'+str(r+3)]="Left"
    sheet['C'+str(r+3)].fill=PatternFill(patternType='solid',fgColor=color)
    sheet['D'+str(r+3)]="Thru"
    sheet['D'+str(r+3)].fill=PatternFill(patternType='solid',fgColor=color)
    sheet['E'+str(r+3)]="Right"
    sheet['E'+str(r+3)].fill=PatternFill(patternType='solid',fgColor=color)
    sheet['F'+str(r+3)]="Left"
    sheet['F'+str(r+3)].fill=PatternFill(patternType='solid',fgColor=color)
    sheet['G'+str(r+3)]="Thru"
    sheet['G'+str(r+3)].fill=PatternFill(patternType='solid',fgColor=color)
    sheet['H'+str(r+3)]="Right"
    sheet['H'+str(r+3)].fill=PatternFill(patternType='solid',fgColor=color)
    sheet['I'+str(r+3)]="Left"
    sheet['I'+str(r+3)].fill=PatternFill(patternType='solid',fgColor=color)
    sheet['J'+str(r+3)]="Thru"
    sheet['J'+str(r+3)].fill=PatternFill(patternType='solid',fgColor=color)
    sheet['K'+str(r+3)]="Right"
    sheet['K'+str(r+3)].fill=PatternFill(patternType='solid',fgColor=color)
    sheet['L'+str(r+3)]="Left"
    sheet['L'+str(r+3)].fill=PatternFill(patternType='solid',fgColor=color)
    sheet['M'+str(r+3)]="Thru"
    sheet['M'+str(r+3)].fill=PatternFill(patternType='solid',fgColor=color)
    sheet['N'+str(r+3)]="Right"
    sheet['N'+str(r+3)].fill=PatternFill(patternType='solid',fgColor=color)

    
    
#export excel file.................................................................................................................

def exportCSV ():
    export_file_path = filedialog.asksaveasfilename(defaultextension='.xlsx')
    book.save(export_file_path)


#Output in dialogbox...................................................................................................................
def output():
    root= tk.Tk()
    canvas1 = tk.Canvas(root, width = 300, height = 200, bg = 'lightsteelblue2', relief = 'raised')
    canvas1.pack()

    saveAsButton_CSV = tk.Button(text='Export LOS Table', command=exportCSV, bg='green', fg='white', font=('helvetica', 12, 'bold'))
    canvas1.create_window(150, 100, window=saveAsButton_CSV)

    root.mainloop()
        


#Program...................................................................................................................................................

check()
output()

